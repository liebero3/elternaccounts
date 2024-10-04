"""
data_processing.py

This module provides functions to process and update data files related to parental accounts.
It focuses on updating XLSX files with data from CSV files and creating parental accounts based 
on form data and Schild CSV exports, adhering to specific data merging and validation rules.

Functions:
- update_xlsx: Updates an XLSX file with data from a CSV file and backs it up.
- createElternaccounts: Matches and generates parental accounts using form data and Schild CSV exports.

Requirements:
- pandas: For data manipulation.
- openpyxl: For handling XLSX file operations.
- Additional custom modules: file_operations, elternaccounts_credentials, mappings, utils.
- Logging is configured to capture debug information.

Note:
Credentials and Nextcloud paths are managed through the elternaccounts_credentials module.
"""
from datetime import datetime
import pandas as pd
import openpyxl as px
from file_operations import put_file
import elternaccounts_credentials
import mappings
from utils import similar, returnUsername
import logging

logger = logging.getLogger(__name__)


NEXTCLOUD_USERNAME = elternaccounts_credentials.username
NEXTCLOUD_PASSWORD = elternaccounts_credentials.password
webdav_backup = elternaccounts_credentials.url_elternaccounts_backup
webdav_share = elternaccounts_credentials.url_elternaccounts_share


def update_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Update an XLSX file with data from a CSV file.

    This function reads data from a specified CSV file and updates an existing XLSX file by merging and 
    removing duplicate timestamp entries. It also adjusts the column widths of the XLSX file for better 
    readability, and creates a backup of the original XLSX file in a Nextcloud directory.

    Args:
        csv_path (str): The path to the source CSV file.
        xlsx_path (str): The path to the target XLSX file to be updated.

    Returns:
        None

    Side Effects:
        - Modifies the specified XLSX file.
        - Saves a backup of the original XLSX file in a Nextcloud directory.

    Raises:
        FileNotFoundError: If the specified CSV or XLSX files are not found.
        PermissionError: If the script does not have write permissions to the XLSX file.
    """
    now = datetime.now()
    date_string = now.strftime("%y%m%d%H%M%S")
    csv_df = pd.read_csv(csv_path)
    xlsx_df = pd.read_excel(xlsx_path)

    backup_url = f"{webdav_backup[:-5]}_backup{date_string}.xlsx"
    put_file(
        backup_url,
        xlsx_path,
        NEXTCLOUD_USERNAME,
        NEXTCLOUD_PASSWORD,
    )

    csv_df["Kontrolliert"] = pd.NA
    merged_df = pd.concat([xlsx_df, csv_df], ignore_index=True)
    merged_df.drop_duplicates(subset="Zeitstempel", keep="first", inplace=True)
    merged_df.to_excel(xlsx_path, index=False)

    wb = px.load_workbook(xlsx_path)
    ws = wb.active
    ws.auto_filter.ref = ws.dimensions

    # Anpassung der Spaltenbreiten
    for column in ["F", "P", "I", "L", "O"]:
        max_length = (
            max(len(str(cell.value)) if cell.value else 0 for cell in ws[column]) + 2
        )
        ws.column_dimensions[column].width = max_length

    # Setzen spezifischer Spaltenbreiten
    spaltenbreiten = {
        "A": 2,
        "B": 2,
        "C": 2,
        "D": 9,
        "E": 15,
        "G": 15,
        "H": 9,
        "J": 9,
        "K": 9,
        "M": 9,
        "N": 9,
        "Q": 11,
    }

    for spalte, breite in spaltenbreiten.items():
        ws.column_dimensions[px.utils.get_column_letter(ws[spalte][0].column)].width = (
            breite
        )

    wb.save(xlsx_path)


def createElternaccounts(
    formsdatei: str, schildexport: str, kontrolloutput: str, outputfile: str
) -> None:
    """
    Create parental accounts based on form data and Schild CSV export.

    This function processes data from a form XLSX file and a Schild CSV export to generate a CSV file of 
    parental accounts. It uses a similarity score to match data from both files. The final results are 
    stored in an output CSV, along with a separate control CSV for verification purposes.

    Args:
        formsdatei (str): Path to the forms XLSX file.
        schildexport (str): Path to the Schild CSV export file.
        kontrolloutput (str): Path to the control output CSV file.
        outputfile (str): Path to the final parental accounts CSV file.

    Returns:
        None

    Side Effects:
        - Writes the control output and final accounts to separate CSV files.

    Raises:
        FileNotFoundError: If any of the specified input files are not found.
        Exception: For other errors during data processing or file writing.
    """
    forms = pd.read_excel(formsdatei)
    schild = pd.read_csv(schildexport, delimiter=";", quotechar='"')

    outputtest = []
    output = []

    forms_filtered = forms[forms["Kontrolliert"] == 1]

    for idx, form_row in forms_filtered.iterrows():
        for i in range(1, 4):
            fname = form_row.get(f"Vorname des {i}. Kindes")
            lname = form_row.get(f"Nachname des {i}. Kindes")
            klasse = form_row.get(f"Klasse des {i}. Kindes")
            if pd.notna(fname):
                #   print(f"{fname} {lname} {klasse}")
                matching_schild = None
                second_matching_schild = None
                highest_similarity = 0
                second_highest_similarity = 0
                for _, schild_row in schild.iterrows():
                    if schild_row["webuntisKlasse"] == klasse:
                        full_name_schild = (
                            f"{schild_row['US_firstName']} {schild_row['US_lastName']}"
                        )
                        full_name_forms = f"{fname} {lname}"
                        similarity = similar(full_name_schild, full_name_forms)
                        if similarity > highest_similarity:
                            second_highest_similarity = highest_similarity
                            highest_similarity = similarity
                            second_matching_schild = matching_schild
                            matching_schild = schild_row
                        elif similarity > second_highest_similarity:
                            second_highest_similarity = similarity
                            second_matching_schild = schild_row

                if matching_schild is not None:
                    if highest_similarity <= 0.9 and second_matching_schild is not None:
                        logger.debug(
                            f"Elternteil: {form_row['Vorname des Elternteils']} {form_row['Nachname des Elternteils']}, "
                            f"Kind (Forms): {fname} {lname}, "
                            f"Matching Kind (Schild): {matching_schild['US_firstName']} {matching_schild['US_lastName']}, "
                            f"Student-ID: {matching_schild['AT_webuntisUid']}, "
                            f"Similarity Score: {highest_similarity:.2f}, "
                            f"Zweites Matching Kind (Schild): {second_matching_schild['US_firstName']} {second_matching_schild['US_lastName']}, "
                            f"Student-ID: {second_matching_schild['AT_webuntisUid']}, "
                            f"Second Highest Similarity: {second_highest_similarity:.2f}"
                        )
                        logger.debug(
                            f"{form_row['Vorname des Elternteils']};"
                            f"{form_row['Nachname des Elternteils']};"
                            f"{form_row['Emailadresse des Elternteils'].lower()};"
                            f"{matching_schild['AT_webuntisUid']}"
                        )
                    #   if highest_similarity <= 0.9 and second_matching_schild is not None:
                    #   print(
                    #       form_row["Vorname des Elternteils"],
                    #       form_row["Nachname des Elternteils"],
                    #       fname,
                    #       lname,
                    #       matching_schild["US_firstName"],
                    #       matching_schild["US_lastName"],
                    #       matching_schild["AT_webuntisUid"],
                    #       highest_similarity,
                    #       second_matching_schild["US_firstName"],
                    #       second_matching_schild["US_lastName"],
                    #       second_matching_schild["AT_webuntisUid"],
                    #       second_highest_similarity,
                    #   )
                    #   print(
                    #       f'{form_row["Vorname des Elternteils"]};{form_row["Nachname des Elternteils"]};{form_row["Emailadresse des Elternteils"].lower()};{matching_schild["AT_webuntisUid"]}'
                    #   )

                    outputtest.append(
                        [
                            form_row["Vorname des Elternteils"],
                            form_row["Nachname des Elternteils"],
                            form_row["Emailadresse des Elternteils"].lower(),
                            matching_schild["AT_webuntisUid"],
                            fname,
                            lname,
                            matching_schild["US_firstName"],
                            matching_schild["US_lastName"],
                            matching_schild["AT_webuntisUid"],
                            highest_similarity,
                            second_highest_similarity,
                        ]
                    )
                    if highest_similarity > 0.5:
                        output.append(
                            [
                                form_row["Vorname des Elternteils"],
                                form_row["Nachname des Elternteils"],
                                form_row["Emailadresse des Elternteils"].lower(),
                                matching_schild["AT_webuntisUid"],
                            ]
                        )

    output_df = pd.DataFrame(
        outputtest,
        columns=[
            "Eltern Vorname",
            "Eltern Nachname",
            "email",
            "student-id",
            "Kind Vorname (forms)",
            "Kind Nachname (forms)",
            "Kind Vorname (schild)",
            "Kind Nachname (schild)",
            "AT_webuntisUid",
            "Best Similarity Score",
            "Second Best Similarity Score",
        ],
    )
    output_df.to_csv(kontrolloutput, index=False, sep=";")
    output_df2 = pd.DataFrame(
        output, columns=["Eltern Vorname", "Eltern Nachname", "email", "student-id"]
    )
    output_df2["username"] = output_df2.apply(
        lambda row: returnUsername(
            row["Eltern Vorname"], row["Eltern Nachname"], "kurzform"
        ),
        axis=1,
    )
    output_df2.to_csv(outputfile, index=False, sep=";")
